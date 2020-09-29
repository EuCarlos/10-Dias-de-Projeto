# encoding: utf-8
from tkinter import *
from openpyxl import load_workbook
from datetime import date
from converterArq import Excel2Pdf, apagaArq

arq = open('NomeDosAlunos.txt')
alunos = arq.readlines()

def gerarCertificado():

    for i in range(len(alunos)):
        prof = FProf.get()
        planilha1.cell(row=13, column=2, value=alunos[i])
        planilha1.cell(row=19, column=5, value=date)
        planilha1.cell(row=26, column=2, value=prof)
        arquivo_excel.save('Resultados/Certificado{}.xlsx'.format(i))
        txt4 = Label(jan, text='CERTIFICADOS GERADOS', bg='green', fg='white').place(x=90, y=160)

    for i in range(len(alunos)):
        input_file = r'C:\[caminho_do_arquivo]\Resultados\Certificado{}.xlsx'.format(i)
        output_file = r'C:\[caminho_do_arquivo]\Resultados\Certificado{}.pdf'.format(i)
        Excel2Pdf(input_file, output_file)

    for i in range(len(alunos)):
        input_file = r'C:\[caminho_do_arquivo]\Resultados\Certificado{}.xlsx'.format(i)
        apagaArq(input_file)

date_today = date.today()
date = '{}/{}/{}'.format(date_today.day, date_today.month, date_today.year)
corBG = '#f2cc8f'
cor1 = '#81b29a'

caminho = 'Modelo.xlsx'
arquivo_excel = load_workbook(caminho)
planilha1 = arquivo_excel.active

jan = Tk()
jan.title('Gerar Certificado')
jan['bg'] = corBG
jan.geometry('300x200+500+300')
jan.resizable(width=False, height=False)

txt1 = Label(jan, text='GERAR CERTIFICADO', bg=cor1, fg='black').pack(side=TOP, fill=X)
txt2 = Label(jan, text='DATA DOS CERTIFICADOS: '+ date, bg=corBG, fg='black').place(x=20, y=50)
txt3 = Label(jan, text='NOME DO INTRUTOR:', bg=corBG, fg='black').place(x=20, y=85)
FProf = Entry(jan)
FProf.place(x=155, y=85)
btnGerar = Button(jan, width='20', text='GERAR CERTIFICADO', fg='white', bg=cor1, command=gerarCertificado).place(x=85, y=120)

jan.mainloop()
